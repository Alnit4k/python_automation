from selenium import webdriver  #permite fazer as automações
from selenium.webdriver.common.keys import Keys  #Permite simular o uso do teclado
from selenium.webdriver.common.by import By  #Permite encontrar elementos
from selenium.webdriver.support.select import Select  #Permite trabalhar com dropdown
from time import sleep
import openpyxl

numero_oab= 133864

driver= webdriver.Chrome()
driver.get('https://pje-consulta-publica.tjmg.jus.br/')

sleep(10) 

campo_oab = driver.find_element(By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)

dropdown_estados = driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
opcoes_estados = Select(dropdown_estados)
opcoes_estados.select_by_visible_text('SP')

botao_pesquisar = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
botao_pesquisar.click ()
sleep(10)

processos = driver.find_elements(By.XPATH,"//b[@class='btn-block']")
for processo in processos:
    processo.click()
    sleep(10)
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])
    driver.set_window_size(1920,1080)
    numero_processo = driver.find_elements(By.XPATH, "//div[@class='col-sm-12']")
    numero_processo = numero_processo[0]
    numero_processo = numero_processo.text
    

    data_distribuicao = driver._elements(By.XPATH, "//div[@class='value col-sm-12']")
    data_distribuicao = data_distribuicao[1]
    data_distribuicao = data_distribuicao.text
    # extrair e guardar todas as ultimas movimentações
    movimentacoes = driver.find_elements(By.XPATH, "//div[@id=j_id132:processoEventoPanel_body]//tr[cotains(@class,'rich-table-row')]//td//div//div//span")
    lista_movimentacoes = []

    for movimentacao in movimentacoes:
        lista_movimentacoes.append(movimentacao.text)
    #guardar tudo no exel, separados por processo
    workbook = openpyxl.load_workbook('dados.xlsx')
    try:
        #codigo para inserir dados em página existente
        # acessar pagina do processo
        pagina_processo = workbook[numero_processo]
        # criar nome das colunas
        pagina_processo['A1'].value = "Número do processo"
        pagina_processo['B1'].value = "Data Distribuição"
        pagina_processo['C1'].value = "Movimentações"
        # adicionar número do processo
        pagina_processo['A2'].value = numero_processo
        # adicionar data de distribuição        
        pagina_processo['B2'].value = data_distribuicao 
        # adicionar movimentações
        for index, movimentacao in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), min_col=3, max_col=3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save('dados.xlsx')
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
    except Exception as error:
        # código para criar uma página do zero e inserir as informações
        workbook.create_sheet(numero_processo)
        pagina_processo = workbook[numero_processo]
        # criar nome das colunas
        pagina_processo['A1'].value = "Número do processo"
        pagina_processo['B1'].value = "Data Distribuição"
        pagina_processo['C1'].value = "Movimentações"
        # adicionar número do processo
        pagina_processo['A2'].value = numero_processo
        # adicionar data de distribuição        
        pagina_processo['B2'].value = data_distribuicao 
        # adicionar movimentações
        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), min_col=3, max_col=3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]
        workbook.save('dados.xlsx')
        driver.close()
        driver.switch_to.window(driver.window_handles[0])